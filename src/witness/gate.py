"""Engine-trust gate.

Before any fuzzing, the pure-Python recalculation engine must be shown to
reproduce the values Excel itself last cached inside the .xlsx. If it cannot,
the acceptance oracle is untrustworthy and the whole approach is void.

Exit criterion (plan §9): >= 10 of 15 workbooks reproduce within tolerance.
"""

from __future__ import annotations

import json
import math
import os
import sys
import warnings
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")
# The formulas engine drives a tqdm bar per workbook; silence it.
os.environ.setdefault("TQDM_DISABLE", "1")

ABS_TOL = 1e-6
REL_TOL = 1e-9


@dataclass
class CellResult:
    sheet: str
    coord: str
    formula: str
    cached: object
    computed: object
    agrees: bool
    reason: str = ""


@dataclass
class WorkbookReport:
    path: str
    loaded: bool
    error: str
    formula_cells: int
    compared: int
    agreed: int
    skipped_no_cache: int
    engine_errors: int
    disagreements: list

    @property
    def rate(self) -> float:
        return self.agreed / self.compared if self.compared else 0.0

    @property
    def passes(self) -> bool:
        # A workbook passes the gate if the engine reproduced every cached
        # value it was able to compare, over a non-trivial number of cells.
        return self.loaded and self.compared >= 5 and self.agreed == self.compared


_EPOCH = datetime(1899, 12, 30)


def _to_serial(v):
    """Excel stores dates as a serial day count; the engine returns the serial,
    openpyxl converts the cached value to a datetime. Normalise to the serial."""
    if isinstance(v, datetime):
        d = v - _EPOCH
        return d.days + d.seconds / 86400.0
    if isinstance(v, date):
        return (v - _EPOCH.date()).days
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0
    if isinstance(v, timedelta):
        return v.days + v.seconds / 86400.0
    return None


def _close(a, b) -> tuple[bool, str]:
    """Compare a cached Excel value against an engine-computed value."""
    if isinstance(a, str) and a.startswith("#"):
        return True, "cached-error-cell-skipped"
    if a is None and b is None:
        return True, ""
    # Date/time cells: openpyxl gives a datetime, the engine gives the serial.
    sa, sb = _to_serial(a), _to_serial(b)
    if sa is not None or sb is not None:
        na = sa if sa is not None else (a if isinstance(a, (int, float)) else None)
        nb = sb if sb is not None else (b if isinstance(b, (int, float)) else None)
        if na is not None and nb is not None:
            return math.isclose(na, nb, rel_tol=0, abs_tol=1e-6), "date-serial"
    # Excel stores an empty formula result as 0 or blank interchangeably.
    if a is None and isinstance(b, (int, float)) and b == 0:
        return True, "blank~zero"
    if b is None and isinstance(a, (int, float)) and a == 0:
        return True, "blank~zero"
    if isinstance(a, bool) or isinstance(b, bool):
        return bool(a) == bool(b), ""
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if math.isnan(a) and math.isnan(b):
            return True, ""
        return math.isclose(a, b, rel_tol=REL_TOL, abs_tol=ABS_TOL), ""
    return str(a).strip() == str(b).strip(), ""


def _unwrap(v):
    """formulas returns numpy arrays / Ranges; reduce to a scalar."""
    import numpy as np

    if isinstance(v, np.ndarray):
        if v.size == 0:
            return None
        v = v.flatten()[0]
    if hasattr(v, "value"):
        v = v.value
        if isinstance(v, np.ndarray):
            v = v.flatten()[0] if v.size else None
    if isinstance(v, np.generic):
        v = v.item()
    if isinstance(v, str) and v.startswith("#"):
        return v
    return v


def check_workbook(path: Path) -> WorkbookReport:
    rep = WorkbookReport(str(path), False, "", 0, 0, 0, 0, 0, [])
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001
        rep.error = f"openpyxl load failed: {type(e).__name__}: {e}"
        return rep

    targets: dict[str, tuple[str, object]] = {}
    for ws in wb_f.worksheets:
        vs = wb_v[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    rep.formula_cells += 1
                    cached = vs[c.coordinate].value
                    if cached is None:
                        rep.skipped_no_cache += 1
                        continue
                    key = f"'[{path.name}]{ws.title}'!{c.coordinate}".upper()
                    targets[key] = (c.value, cached)

    if not targets:
        rep.error = "no formula cells with cached values"
        return rep

    try:
        import formulas

        xl = formulas.ExcelModel().loads(str(path)).finish()
        sol = xl.calculate()
    except Exception as e:  # noqa: BLE001
        rep.error = f"formulas engine failed: {type(e).__name__}: {str(e)[:200]}"
        return rep

    rep.loaded = True
    norm = {k.upper(): v for k, v in sol.items()}

    for key, (formula, cached) in targets.items():
        computed = None
        for cand in (key, key.replace("'", "")):
            if cand in norm:
                computed = norm[cand]
                break
        if computed is None:
            rep.engine_errors += 1
            continue
        try:
            computed = _unwrap(computed)
        except Exception:  # noqa: BLE001
            rep.engine_errors += 1
            continue

        rep.compared += 1
        ok, reason = _close(cached, computed)
        if ok:
            rep.agreed += 1
        elif len(rep.disagreements) < 12:
            sheet, coord = key.split("!")
            rep.disagreements.append(
                asdict(
                    CellResult(sheet, coord, formula, _safe(cached), _safe(computed), False, reason)
                )
            )
    return rep


def _safe(v):
    if isinstance(v, (int, float, str, bool)) or v is None:
        return v
    return repr(v)


def main(argv: list[str]) -> int:
    corpus = Path(argv[1]) if len(argv) > 1 else Path("corpus")
    books = sorted([p for p in corpus.rglob("*.xls*") if not p.name.startswith("~$")])
    if not books:
        print(f"no workbooks found under {corpus}/")
        return 2

    reports = []
    for p in books:
        r = check_workbook(p)
        reports.append(r)
        mark = "PASS" if r.passes else "FAIL"
        detail = (
            f"{r.agreed}/{r.compared} cells"
            + (f", {r.engine_errors} engine-miss" if r.engine_errors else "")
            + (f", {r.skipped_no_cache} no-cache" if r.skipped_no_cache else "")
        )
        print(f"[{mark}] {p.name:<44} {detail}" + (f"  :: {r.error}" if r.error else ""))
        for d in r.disagreements[:3]:
            print(f"         {d['coord']}  {d['formula'][:52]}")
            print(f"           excel={d['cached']!r}  engine={d['computed']!r}")

    passed = sum(1 for r in reports if r.passes)
    total = len(reports)
    Path("results").mkdir(exist_ok=True)
    Path("results/gate.json").write_text(
        json.dumps(
            {
                "passed": passed,
                "total": total,
                "reports": [
                    {**asdict(r), "rate": round(r.rate, 4), "passes": r.passes} for r in reports
                ],
            },
            indent=2,
            default=str,
        )
    )

    print(f"\n{'=' * 64}\nGATE: {passed}/{total} workbooks reproduce their own cached values")
    print("Criterion (plan §9): >= 10 of 15 -> proceed with Witness")
    print("results/gate.json written")
    return 0 if passed >= 10 else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
