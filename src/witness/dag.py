"""Formula-DAG extraction.

Deterministic. No model is involved at any point in this module — that is the
entire reason it exists. An LLM asked to port a workbook cannot reliably tell a
*true input* from a *derived cell*, and porting a derived cell as an input is a
whole family of silent failures. The dependency graph cannot make that mistake.

The other job here is typing each input's domain. The fuzzer is only as good as
the values it samples, and a uniform random float finds nothing. Lifting tier
boundaries out of the lookup tables a formula actually references is what makes
the fuzzer find tier-boundary bugs instead of noise.
"""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

warnings.filterwarnings("ignore")

# 'Sheet Name'!$A$1  |  SheetName!A1:B9  |  A1  |  $A$1:$C$3
_REF = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(?P<ref>\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)"
)
_FUNC = re.compile(r"([A-Z][A-Z0-9._]*)\s*\(")
_STRINGS = re.compile(r'"[^"]*"')

# Functions whose presence means a cell's value depends on something the
# recalculation engine cannot be trusted to reproduce deterministically.
NONDETERMINISTIC = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT"}


@dataclass
class Cell:
    key: str
    sheet: str
    coord: str
    formula: str | None
    value: object
    precedents: set[str] = field(default_factory=set)
    dependents: set[str] = field(default_factory=set)
    functions: set[str] = field(default_factory=set)

    @property
    def is_formula(self) -> bool:
        return self.formula is not None


@dataclass
class InputSpec:
    """A true input cell, plus the domain the fuzzer should sample from."""

    key: str
    kind: str  # "number" | "text" | "bool" | "date" | "blank"
    observed: object
    interesting: list = field(default_factory=list)
    used_by_functions: set[str] = field(default_factory=set)

    def to_dict(self) -> dict:
        return {
            "key": self.key,
            "kind": self.kind,
            "observed": _plain(self.observed),
            "interesting": [_plain(v) for v in self.interesting],
            "used_by": sorted(self.used_by_functions),
        }


@dataclass
class WorkbookDAG:
    path: str
    cells: dict[str, Cell]
    inputs: list[InputSpec]
    outputs: list[str]
    nondeterministic: list[str]

    @property
    def formula_cells(self) -> list[Cell]:
        return [c for c in self.cells.values() if c.is_formula]

    def summary(self) -> dict:
        return {
            "path": self.path,
            "cells": len(self.cells),
            "formula_cells": len(self.formula_cells),
            "inputs": len(self.inputs),
            "outputs": len(self.outputs),
            "nondeterministic": self.nondeterministic,
        }


def _plain(v):
    if isinstance(v, (int, float, str, bool)) or v is None:
        return v
    return str(v)


def _key(sheet: str, coord: str) -> str:
    return f"{sheet}!{coord.replace('$', '').upper()}"


def _expand(sheet: str, ref: str) -> list[str]:
    """Expand A1 or A1:C3 into individual cell keys."""
    ref = ref.replace("$", "")
    try:
        min_c, min_r, max_c, max_r = range_boundaries(ref)
    except Exception:  # noqa: BLE001
        return []
    if None in (min_c, min_r, max_c, max_r):
        return []
    # Guard against whole-column references blowing up memory.
    if (max_r - min_r) * (max_c - min_c) > 100_000:
        return []
    return [
        _key(sheet, f"{get_column_letter(c)}{r}")
        for r in range(min_r, max_r + 1)
        for c in range(min_c, max_c + 1)
    ]


def _parse_refs(formula: str, home_sheet: str) -> list[str]:
    body = _STRINGS.sub('""', formula)
    out: list[str] = []
    for m in _REF.finditer(body):
        sheet = m.group("sheet") or home_sheet
        sheet = sheet.strip("'")
        out.extend(_expand(sheet, m.group("ref")))
    return out


def _lookup_boundaries(cells: dict[str, Cell], formula: str, home_sheet: str) -> list:
    """For VLOOKUP/HLOOKUP/MATCH with approximate match, the tier boundaries in
    the referenced table are the values most likely to break a port."""
    vals = []
    for rng in _parse_refs(formula, home_sheet):
        c = cells.get(rng)
        if c and not c.is_formula and isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
            vals.append(c.value)
    uniq = sorted(set(vals))[:24]
    # The boundary itself, and the values immediately either side of it.
    edge = []
    for v in uniq:
        edge.extend([v, v - 0.01, v + 0.01])
    return edge[:48]


def build(path: Path) -> WorkbookDAG:
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)

    cells: dict[str, Cell] = {}
    for ws in wb_f.worksheets:
        vs = wb_v[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                is_f = isinstance(c.value, str) and c.value.startswith("=")
                k = _key(ws.title, c.coordinate)
                cells[k] = Cell(
                    key=k,
                    sheet=ws.title,
                    coord=c.coordinate,
                    formula=c.value if is_f else None,
                    value=vs[c.coordinate].value if is_f else c.value,
                )

    # Edges.
    nondet: list[str] = []
    for c in list(cells.values()):
        if not c.is_formula:
            continue
        c.functions = set(_FUNC.findall(c.formula.upper()))
        if c.functions & NONDETERMINISTIC:
            nondet.append(c.key)
        for ref in _parse_refs(c.formula, c.sheet):
            if ref == c.key:
                continue
            c.precedents.add(ref)
            if ref in cells:
                cells[ref].dependents.add(c.key)

    # True inputs: referenced by at least one formula, and not a formula itself.
    referenced: set[str] = set()
    for c in cells.values():
        if c.is_formula:
            referenced |= c.precedents

    inputs: list[InputSpec] = []
    for k in sorted(referenced):
        c = cells.get(k)
        if c is None:
            # Referenced but empty — still a real input the fuzzer can fill.
            inputs.append(InputSpec(key=k, kind="blank", observed=None, interesting=[0, "", None]))
            continue
        if c.is_formula:
            continue
        v = c.value
        if isinstance(v, bool):
            kind, interesting = "bool", [True, False]
        elif isinstance(v, (int, float)):
            kind = "number"
            interesting = [0, -abs(v) if v else -1, v, v * 2 if v else 1, None, ""]
        elif isinstance(v, str):
            kind, interesting = "text", ["", v, "0"]
        else:
            kind, interesting = "date", [v]
        spec = InputSpec(key=k, kind=kind, observed=v, interesting=list(interesting))
        for dep in c.dependents:
            spec.used_by_functions |= cells[dep].functions
        inputs.append(spec)

    # Tier boundaries from approximate-match lookups.
    by_key = {i.key: i for i in inputs}
    for c in cells.values():
        if not c.is_formula:
            continue
        if not (c.functions & {"VLOOKUP", "HLOOKUP", "MATCH", "LOOKUP"}):
            continue
        if not re.search(r",\s*(TRUE|1)\s*\)", c.formula.upper()):
            continue
        bounds = _lookup_boundaries(cells, c.formula, c.sheet)
        for ref in c.precedents:
            s = by_key.get(ref)
            if s and s.kind == "number":
                s.interesting = list(dict.fromkeys(s.interesting + bounds))[:64]

    outputs = [c.key for c in cells.values() if c.is_formula and not c.dependents]

    return WorkbookDAG(str(path), cells, inputs, sorted(outputs), sorted(nondet))


def slice_for_output(dag: WorkbookDAG, target: str, max_inputs: int = 40) -> dict:
    """Transitive input closure of one output cell.

    A whole workbook can carry 195,000 input cells, which is not a fuzzable
    surface and is not what the user cares about anyway. Owen cares about one
    number — the recognised revenue, the commission total. Scoping a case to a
    single output and everything upstream of it makes the domain small enough to
    sample densely, and makes the certificate say something a human can act on.
    """
    seen: set[str] = set()
    stack = [target]
    inputs: list[str] = []
    depth = 0
    while stack:
        k = stack.pop()
        if k in seen:
            continue
        seen.add(k)
        c = dag.cells.get(k)
        if c is None:
            inputs.append(k)
            continue
        if not c.is_formula:
            inputs.append(k)
            continue
        depth += 1
        stack.extend(c.precedents)

    by_key = {i.key: i for i in dag.inputs}
    specs = [by_key[k] for k in inputs if k in by_key]
    # Prefer inputs feeding lookups (tier boundaries) — they break ports most.
    specs.sort(key=lambda s: (not (s.used_by_functions & {"VLOOKUP", "HLOOKUP", "MATCH", "LOOKUP"}), s.key))
    nondet = sorted(set(dag.nondeterministic) & seen)
    return {
        "target": target,
        "formula_nodes": depth,
        "inputs": specs[:max_inputs],
        "truncated": len(specs) > max_inputs,
        "total_inputs": len(specs),
        "nondeterministic": nondet,
    }


def pick_cases(dag: WorkbookDAG, n: int = 3, min_depth: int = 3, max_inputs: int = 60) -> list[dict]:
    """Choose the most interesting output cells to certify in this workbook."""
    cands = []
    for out in dag.outputs:
        sl = slice_for_output(dag, out, max_inputs=max_inputs)
        if sl["nondeterministic"]:
            continue  # cannot be an oracle: value depends on NOW()/RAND()/etc.
        if sl["formula_nodes"] < min_depth or not sl["inputs"]:
            continue
        if sl["truncated"]:
            continue
        cands.append(sl)
    # Deepest dependency chains first — the most logic behind one number.
    cands.sort(key=lambda s: (-s["formula_nodes"], -len(s["inputs"])))
    picked, used_sheets = [], {}
    for c in cands:
        sheet = c["target"].split("!")[0]
        if used_sheets.get(sheet, 0) >= 5:
            continue  # spread across sheets rather than many cells in one column
        used_sheets[sheet] = used_sheets.get(sheet, 0) + 1
        picked.append(c)
        if len(picked) >= n:
            break
    return picked


if __name__ == "__main__":
    import json
    import sys

    corpus = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("corpus")
    rows = []
    for p in sorted(corpus.glob("*.xlsx")):
        try:
            d = build(p)
        except Exception as e:  # noqa: BLE001
            print(f"[ERR ] {p.name}: {type(e).__name__}: {str(e)[:90]}")
            continue
        s = d.summary()
        rows.append({**s, "input_specs": [i.to_dict() for i in d.inputs[:400]]})
        nd = f"  NONDET:{len(d.nondeterministic)}" if d.nondeterministic else ""
        print(
            f"[ OK ] {p.name:<46} {s['formula_cells']:>6} formulas "
            f"{s['inputs']:>5} inputs {s['outputs']:>5} outputs{nd}"
        )
    Path("results").mkdir(exist_ok=True)
    Path("results/dag.json").write_text(json.dumps(rows, indent=2, default=str))
    print(f"\n{len(rows)} workbooks -> results/dag.json")
